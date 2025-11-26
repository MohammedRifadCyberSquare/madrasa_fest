from django.shortcuts import render, redirect
from .models import MadrasaAdmin, Participant, ParticipantItem
from openpyxl import load_workbook
from django.db.models import Max
from .constants import CATEGORIES
from django.http import JsonResponse
from .utils import render_pdf
import os
from django.views.decorators.clickjacking import xframe_options_exempt
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings


def login(request):
    if request.method == 'POST':
        username = request.POST['user_name']
        password = request.POST['password']
        try:
            admin = MadrasaAdmin.objects.get(
                user_name=username, password=password)
            request.session['admin_id'] = admin.id
            return redirect('dashboard')
        except MadrasaAdmin.DoesNotExist:
            return render(request, 'fest/login.html', {'error': 'UserName or password Incorrect!'})

    return render(request, 'fest/login.html')


def dashboard(request):
    return render(request, 'fest/home.html')


def bulk_registration(request):
    if request.method == 'POST':
        file = request.FILES['file']
        try:
            workbook = load_workbook(filename=file)
            sheet = workbook.active
            rows_added = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if all(cell is None for cell in row):
                    continue

                row = list(row[1:])  # skip first column if needed
                madrasa_name, cls_madrasa, student_name, father_name, phone_no, \
                    dob, gender, category, house_name, items = row[:10]

                # Determine chest_no
                last_chest = Participant.objects.filter(house_name__iexact=house_name)\
                    .aggregate(Max('chest_no'))['chest_no__max']

                if not last_chest:
                    chest_no = {'blue': 200, 'red': 400, 'green': 600,
                                'yellow': 800}.get(house_name.lower(), 0)
                else:
                    chest_no = last_chest + 1

                # Create participant
                participant = Participant.objects.create(
                    madrasa_name=madrasa_name.lower(),
                    cls_madrasa=int(cls_madrasa),
                    student_name=student_name.lower(),
                    father_name=father_name.lower(),
                    phone_no=str(phone_no),
                    dob=str(dob),
                    gender=gender.lower(),
                    category=category.lower(),
                    house_name=house_name.lower(),
                    chest_no=chest_no,
                )

                # Add items
                if items:
                    item_list = [i.strip() for i in items.split(',')]
                    for item_name in item_list:
                        ParticipantItem.objects.create(
                            participant=participant,
                            item=item_name.lower()
                        )

                rows_added += 1  # increment inside the loop

            # messages.success(request, f"{rows_added} participants added successfully!")
            return redirect('bulk_upload')

        except Exception as e:
            pass
            # mes
            # passsages.error(request, f"Error: {e}")

    return render(request, 'fest/bulk_registration.html')


def get_events(request):
    category = request.GET.get('category', '')   # selected category name
    events = CATEGORIES.get(category, [])        # get items in that category
    return JsonResponse({'events': events})


def calling_list(request):
    participants = Participant.objects.all().order_by('house_name', 'chest_no')
    return render(request, 'fest/calling_list.html', {
        'participants': participants,
        'categories': CATEGORIES
    })


@csrf_exempt
@xframe_options_exempt
def generate_calling_list(request):
    if request.method == "POST":
        category = request.POST.get("category")
        item = request.POST.get("item")
        report_type = request.POST.get("report_type", "calling_list",)

        if report_type == "calling_list":
            template_path = "report_templates/call_list_template.html"
        elif report_type == "green_room":
            template_path = "report_templates/green_room_template.html"
        else:
            template_path = "report_templates/valuation_template.html"
        if not category or not item:
            return JsonResponse({"status": "error", "message": "Category and item required"})

        participants = Participant.objects.filter(
            category=category.lower(),
            items__item=item.lower()
        ).order_by('house_name', 'chest_no').distinct()

        extra_rows = range(max(0, 10 - len(participants))
                           ) if report_type == "green_room" or "valuation" else []

        filename = f"{report_type}_{category}_{item}.pdf"
        logo_path = os.path.join(
            settings.BASE_DIR, "static", "images", "logo1.jpg")

        pdf_path = render_pdf(
            template_path,
            {
                "category": category,
                "item": item,
                "participants": participants,
                "logo_path": logo_path,
                "extra_rows": extra_rows,
            },
            filename
        )

        print(pdf_path, '++++++++++++++')
        safe_filename = filename.replace(" ", "_")
        pdf_url = settings.MEDIA_URL + "pdfs/" + safe_filename
        print(pdf_url, '-------------------------------------------')
        return JsonResponse({"status": "success", "pdf_url": pdf_url})

    return JsonResponse({"status": "error", "message": "Invalid request"})
